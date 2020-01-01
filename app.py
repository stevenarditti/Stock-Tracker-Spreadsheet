#!/usr/bin/env python3

from openpyxl import Workbook
from stock import Stock
import sys
import json


######### Variables
companies = []
data_fields = {}
stocks = set()
src_file = "companies.txt"
save_file = "companies.xlsx"
wb = ""
cur_row = 1
cur_col = ""

# True to print debug info, False to ignore
DEBUG = True

# check if a letter needs to be appended for incrementation, e.g. column ZZ
# returns new col
def inc_col():
    global cur_col
    new_col = ""
    append = True
    chars = list(cur_col)
    for ch in chars:
        if (ch != "Z"):
            append = False
            break
        new_col += "A"
    if (append):
        cur_col = new_col + "A"
    else:
        cur_col = chr(ord(cur_col) + 1)
        return cur_col

def reset_cur_col():
    global cur_col
    cur_col = "A"

def write_company_data():
    global cur_row, cur_col
    ws = wb.active
    for stock in stocks:
        reset_cur_col()
        cur_row += 1

        # set ticker
        ws[cur_col + str(cur_row)] = stock.getTicker()

        data = stock.getData()
        # set data
        for field in data_fields:
            if DEBUG:
                print(field)
                print(data_fields[field])
                print(eval("data" + data_fields[field]))
            ws[inc_col() + str(cur_row)] = eval("data" + data_fields[field])
    wb.save(save_file)

def write_companies():
    for stock in stocks:
        write_company_data()

# util function, consider making a util class?
def list_to_json(read_lines):
    string = ""
    for line in read_lines:
        string += line
    return json.loads(string)

def init_workbook():
    global companies, data_fields, wb

    f1 = open(src_file, "r+")
    companies = f1.readlines()
    f1.close()

    f2 = open("data_mapping.json", "r+")
    data_fields = list_to_json(f2.readlines())
    f2.close()

    wb = Workbook()
    ws = wb.active
    reset_cur_col()
    ws["A1"] = "Ticker"
    for key in data_fields:
        ws[inc_col() + "1"] = key
    wb.save(save_file)
    
    for ticker in companies:
        stocks.add(Stock(ticker))

if __name__ == "__main__":
    ii = 0
    while len(sys.argv) > ii:
        if sys.argv[ii] == "-p" and len(sys.argv) > ii + 1:
            src_file = sys.argv[ii+1]
            try:
                save_file = src_file.split(".")[0] + ".xlsx"
            except:
                save_file = src_file + ".xlsx" 
        ii+=1
    
    init_workbook()
    write_companies()
